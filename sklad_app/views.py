from django.contrib.auth.decorators import login_required
from django.forms import modelformset_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db.models import Count, Sum
from .forms import *
from .models import Pile, NamePile, BrigadeWork, ReturnPile, Car
from django.http import JsonResponse, HttpResponseRedirect,HttpResponse
import json
from django.utils import timezone
from django.shortcuts import render
from .models import *
import re
from django.http import JsonResponse, HttpResponseRedirect
from collections import defaultdict
from django.http import JsonResponse
from .models import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.db.models import F
from datetime import timedelta,datetime
def main(request):
    context = {}
    if request.user.is_authenticated:
        if request.user.userprofile.role.name == "Кладовщик":
            latest_arrivals = OperationArrival.objects.all().order_by('-date')[:10]
            for arrival in latest_arrivals:
                if arrival.details:
                    details = json.loads(arrival.details)
                    arrival.piles_info = []
                    for detail in details:
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            arrival.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                            })
                        except Pile.DoesNotExist:
                            continue
            latest_departures = OperationDeparture.objects.filter(confirm_b=False,visible=True).order_by('-date')[:10]
            for departure in latest_departures:
                if departure.details:
                    details = json.loads(departure.details)
                    departure.piles_info = []
                    for detail in details:
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            departure.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                            })
                        except Pile.DoesNotExist:
                            continue
            latest_opp = SendDetail.objects.filter(confirm=False)[:10]
            for arrival in latest_opp:
                if arrival.details:
                    detail = json.loads(arrival.details)
                    arrival.piles_info = []

                    try:
                        pile = Pile.objects.get(id=detail["pile_id"])
                        arrival.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                        })
                    except Pile.DoesNotExist:
                        continue
            last_latest_opp = SendDetail.objects.filter(confirm=True)[:10]
            for arrival in last_latest_opp:
                if arrival.details:
                    detail = json.loads(arrival.details)
                    arrival.piles_info = []

                    try:
                        pile = Pile.objects.get(id=detail["pile_id"])
                        arrival.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                        })
                    except Pile.DoesNotExist:
                        continue
            context = {
                'latest_arrivals': latest_arrivals,
                'latest_departures': latest_departures,
                'latest_opp':latest_opp,
                'last_latest_opp':last_latest_opp
            }
        if request.user.userprofile.role.name == "Охранник":
            departures = OperationDeparture.objects.filter(confirm=False,confirm_b=True,visible=True)
            for departure in departures:
                if departure.details:
                    details = json.loads(departure.details)
                    departure.piles_info = []
                    for detail in details:
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            departure.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                            })
                        except Pile.DoesNotExist:
                            continue
            confirmed = OperationDeparture.objects.filter(confirm=True,confirm_b=True,visible=True)
            for departure in confirmed:
                if departure.details:
                    details = json.loads(departure.details)
                    departure.piles_info = []
                    for detail in details:
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            departure.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                            })
                        except Pile.DoesNotExist:
                            continue
            context = {
                'departures':departures,
                'confirmed':confirmed

            }
        if request.user.userprofile.role.name == "Менеджер":
            departures = OperationDeparture.objects.filter(confirm=False,visible=True)
            for departure in departures:
                if departure.details:
                    details = json.loads(departure.details)
                    departure.piles_info = []
                    for detail in details:
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            departure.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                            })
                        except Pile.DoesNotExist:
                            continue

            if request.method == "POST":
                departure_id = request.POST.get("departure_id")
                if departure_id:
                    departure = get_object_or_404(OperationDeparture, id=departure_id)
                    departure.confirm = True
                    departure.save()
                    return HttpResponseRedirect(reverse('main'))

            latest_confirmed_departures = OperationDeparture.objects.filter(confirm=True).order_by('-date')[:10]
            for departure in latest_confirmed_departures:
                if departure.details:
                    details = json.loads(departure.details)
                    departure.piles_info = []
                    for detail in details:
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            departure.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                            })
                        except Pile.DoesNotExist:
                            continue
            context = {
                'latest_departures': departures,
                'confirmed': latest_confirmed_departures,
            }
        if request.user.userprofile.role.name == "Производство":

            latest_arrivals = OperationArrival.objects.all().order_by('-date')[:10]
            for arrival in latest_arrivals:
                if arrival.details:
                    details = json.loads(arrival.details)
                    price_parts = json.loads(arrival.price_part)
                    arrival.piles_info = []
                    for detail in details:
                        pile_info = next((item for item in price_parts if item["pile_id"] == detail["pile_id"]), None)
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            quantity = detail["quantity"]
                            part_price1 = float(pile_info.get("part_price", 0))+133*quantity
                            part_price = float(pile_info.get("part_price", 0))
                            price_per_unit = round(part_price / quantity, 2) if quantity > 0 else 0

                            arrival.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"],
                                "part_price": part_price1,
                                "price_per_unit": str(price_per_unit)
                            })
                        except Pile.DoesNotExist:
                            continue
            latest_departures = OperationDeparture.objects.all().order_by('-date')[:10]
            for departure in latest_departures:
                if departure.details:
                    details = json.loads(departure.details)
                    departure.piles_info = []
                    for detail in details:
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            departure.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                            })
                        except Pile.DoesNotExist:
                            continue
            today = timezone.now().date()
            try:
                # Получаем последний заказ за сегодня
                order = Order.objects.filter(date__date=today).latest(
                    'date')  # Предполагается, что в модели Order есть поле date
            except Order.DoesNotExist:
                order = None

            if order:
                # Преобразуем данные о сваях из JSON в словарь
                piles_info = json.loads(order.piles_info)
                materials_total = {
                    'total_beton': order.total_beton,
                    'total_wire_3': order.total_wire_3,
                    'total_wire_4': order.total_wire_4,
                    'total_wire_6': order.total_wire_6,
                    'total_armature': order.total_armature
                }
            else:
                piles_info = {}
                materials_total = {}
            context = {
                'latest_arrivals': latest_arrivals,
                'latest_departures': latest_departures,
                'piles_info': piles_info,
                'materials_total': materials_total
            }

        if request.user.userprofile.role.name == "Бригада":
            user_brigade = UserBrigade.objects.filter(user=request.user).first()
            if user_brigade:
                operation_dep_counts = OperationDepCount.objects.filter(
                    operation__brigade=user_brigade.brigade,
                    visible=True,
                    confirm=False
                ).order_by('-operation__date')  # сортировка по дате операции
            else:
                operation_dep_counts = OperationDepCount.objects.none()
            context = {'operation_dep_counts': operation_dep_counts,}
    return render(request, 'main.html', context)


def jb(request):
    context = {}
    if request.user.is_authenticated and request.user.userprofile.role.name == "Производство":
        latest_arrivals = OperationArrival.objects.all().order_by('-date')[:10]
        for arrival in latest_arrivals:
            if arrival.details:
                details = json.loads(arrival.details)
                price_parts = json.loads(arrival.price_part)
                arrival.piles_info = []
                for detail in details:
                    pile_info = next((item for item in price_parts if item["pile_id"] == detail["pile_id"]), None)
                    try:
                        pile = Pile.objects.get(id=detail["pile_id"])
                        quantity = detail["quantity"]
                        if pile.name.name=='Плита':
                            part_price1 = round(float(pile_info.get("part_price", 0)) + 500 * quantity+11.28 * quantity,3)
                            working=500 * quantity
                            materials=quantity*11.28
                        else:
                            match = re.match(r'(\d+\.?\d*)x(\d+)', pile.size.replace(' ', ''))
                            width, length = map(float, match.groups())
                            if int(length)<4999:
                                part_price1 = round(float(pile_info.get("part_price", 0)) + 115 * quantity+11.28 * quantity,3)
                                working=115 * quantity
                            else:
                                part_price1 = round(float(pile_info.get("part_price", 0)) + 230 * quantity+11.28 * quantity,3)
                                working=230 * quantity
                            materials=quantity*11.28
                        part_price = round(float(pile_info.get("part_price", 0)),4)

                        price_per_unit = round(part_price / quantity, 4) if quantity > 0 else 0

                        arrival.piles_info.append({
                            "pile": pile,
                            "quantity": detail["quantity"],
                            "part_price": part_price,
                            "working": working,
                            "materials":materials,
                            "alls": part_price1,
                            "price_per_unit": str(price_per_unit)
                        })
                    except Pile.DoesNotExist:
                        continue
        latest_departures = OperationDeparture.objects.all().order_by('-date')[:10]
        for departure in latest_departures:
            if departure.details:
                details = json.loads(departure.details)
                departure.piles_info = []
                for detail in details:
                    try:
                        pile = Pile.objects.get(id=detail["pile_id"])
                        departure.piles_info.append({
                            "pile": pile,
                            "quantity": detail["quantity"]
                        })
                    except Pile.DoesNotExist:
                        continue
        today = timezone.now().date()
        try:
            # Получаем последний заказ за сегодня
            order = Order.objects.filter(date__date=today).latest(
                'date')  # Предполагается, что в модели Order есть поле date
        except Order.DoesNotExist:
            order = None

        if order:
            # Преобразуем данные о сваях из JSON в словарь
            piles_info = json.loads(order.piles_info)
            materials_total = {
                'total_beton': order.total_beton,
                'total_wire_3': order.total_wire_3,
                'total_wire_4': order.total_wire_4,
                'total_wire_6': order.total_wire_6,
                'total_armature': order.total_armature,
                'total_armature10': order.total_armature10
            }
        else:
            piles_info = {}
            materials_total = {}
        context = {
            'latest_arrivals': latest_arrivals,
            'latest_departures': latest_departures,
            'piles_info': piles_info,
            'materials_total': materials_total
        }
    return render(request, 'jb.html', context)

def vint(request):
    last_ten_arrivals = OperationArrival_V.objects.order_by('-date')[:10]

    for arrival in last_ten_arrivals:
        if arrival.details:
            details_list = json.loads(arrival.details)
            for detail in details_list:
                try:
                    pile = Pile.objects.get(id=detail["pile_id"])
                    detail["pile_name"] = pile.name.name
                    detail["pile_size"] = pile.size
                    detail["quantity"] = detail.get("quantity", 0)
                    pile_price = Decimal(detail.get("pile_price", "0.0"))
                    detail["pile_price"] = pile_price
                    # Вычисление цены за единицу, предполагая, что количество больше 0
                    detail["price_per_unit"] = (pile_price / detail["quantity"]).quantize(Decimal('.01')) if detail["quantity"] > 0 else Decimal("0.0")
                except Pile.DoesNotExist:
                    detail.update({
                        "pile_name": "Unknown",
                        "pile_size": "Unknown",
                        "price_per_unit": Decimal("0.0")
                    })
            arrival.piles_info = details_list
            arrival.total_price = sum(Decimal(d["pile_price"]) for d in details_list)
    today = timezone.now().date()
    today_orders = Order_v.objects.filter(date__date=today)

    for order in today_orders:
        # Обрабатываем piles
        if order.piles:
            piles_json = json.loads(order.piles)
            piles_info = []
            for pile_data in piles_json:
                try:
                    pile = Pile.objects.get(id=pile_data["pile_id"])
                    if pile_data['quantity'] > 0:
                        piles_info.append({
                            'name': pile.name.name,
                            'size': pile.size,
                            'quantity': pile_data['quantity']
                        })
                except Pile.DoesNotExist:
                    continue
            order.piles_info = piles_info

        # Обрабатываем tubes
        if order.tubes:
            tubes = json.loads(order.tubes)
            order.tubes_info = {key: value for key, value in tubes.items() if value > 0}

        # Обрабатываем lopastis
        if order.lopastis:
            lopastis = json.loads(order.lopastis)
            order.lopastis_info = {key: value for key, value in lopastis.items() if value > 0}


    return render(request, 'vint.html',{'arrivals': last_ten_arrivals,'today_orders': today_orders})

#Сваи
def operation_arrival_view(request):
    if request.user.userprofile.role.name == "Производство":
        piles = Pile.objects.filter(name__pile_type='жб')
        brigades = Brigade.objects.all()
        if request.method == 'POST':
            brigade_id = request.POST.get('brigade')
            brigade = Brigade.objects.get(id=brigade_id)

            wirehouse, _ = WirehouseB.objects.get_or_create(id=1)
            # Получаем последние цены материалов
            last_beton = Beton.objects.last()
            last_wire = Wire.objects.last()
            last_wire4 = Wire4.objects.last()
            last_wire6 = Wire6.objects.last()
            last_armature = Armature.objects.last()
            last_armature10 = Armature10.objects.last()
            total_price = 0
            piles_info = []
            price_parts=[]
            tot_b=0
            tot_arm=0
            tot_arm10=0
            tot_w3=0
            tot_w4=0
            tot_w6=0
            for key, value in request.POST.items():
                if key.startswith('pile_') and value:
                    pile_id = key.split('_')[1]
                    pile = Pile.objects.get(pk=pile_id)

                    quantity_plates = int(value)  # Количество плит введенное пользователем

                    # Использование регулярных выражений для извлечения размеров
                    match = re.match(r'(\d+\.?\d*)x(\d+)', pile.size.replace(' ', ''))
                    if match:
                        width, length = map(float, match.groups())
                        if width == 1.5 and length == 3:
                            multiplier = 1
                        elif width == 150:
                            multiplier = 9
                        elif width == 200:
                            multiplier = 7
                        else:
                            multiplier = 0  # На случай, если размер не соответствует ожидаемым
                    else:
                        multiplier = 0  # Если размер не удалось распознать

                    quantity = quantity_plates * multiplier
                    pile.count += quantity
                    pile.save()
                    if quantity > 0:
                        piles_info.append({"pile_id": pile.id, "quantity": quantity})
                        # Рассчитываем стоимость для каждого материала
                        b_count = pile.price_beton * quantity
                        w3_count = pile.price_wire_3 * quantity
                        w4_count = pile.price_wire_4 * quantity
                        w6_count = pile.price_wire_6 * quantity
                        arm_count = pile.price_armature * quantity

                        tot_b+=b_count
                        tot_w3+=w3_count
                        tot_w4+=w4_count
                        tot_w6+=w6_count
                        oper_del=Del_mat.objects.create(total_beton=b_count,total_wire_3=w3_count, total_wire_4=w4_count, total_wire_6=w4_count)
                        if str(pile.name.name)=='ЖБ ф10' or str(pile.name.name)=='Плита':
                            oper_del.total_armature10=arm_count
                            tot_arm10+=arm_count
                        else:
                            oper_del.total_armature=arm_count
                            tot_arm+=arm_count
                        oper_del.save()
                        beton_price = b_count * (last_beton.price if last_beton else 0)
                        wire3_price = w3_count * (last_wire.price if last_wire else 0)
                        wire4_price = w4_count * (last_wire4.price if last_wire4 else 0)
                        wire6_price = w6_count * (last_wire6.price if last_wire6 else 0)
                        if str(pile.name.name)=='ЖБ ф10' or str(pile.name.name)=='Плита':
                            armature_price = arm_count * (last_armature10.price if last_armature else 0)
                        elif str(pile.name.name)=='ЖБ ф8':
                            armature_price = arm_count * (last_armature.price if last_armature else 0)
                        price_parts.append({
                            "pile_id": pile.id,
                            "quantity": quantity,
                            "part_price": str(round(beton_price + wire3_price + wire4_price + wire6_price + armature_price, 4))
                        })
                        # Суммируем стоимости
                        total_price += beton_price + wire3_price + wire4_price + wire6_price + armature_price

            # Обновляем количество материалов на складе
            wirehouse.wire3 -= tot_w3
            wirehouse.wire4 -= tot_w4
            wirehouse.wire6 -= tot_w6
            wirehouse.armature10 -= tot_arm10
            wirehouse.armature -= tot_arm
            wirehouse.beton -= tot_b
            wirehouse.save()

            # Создаем операцию прихода со всей информацией о сваях
            operation_arrival=OperationArrival.objects.create(
                details=json.dumps(piles_info),
                brigade=brigade,
                price=str(round(total_price,4)),
            )
            operation_arrival.price_part = json.dumps(price_parts)
            operation_arrival.user.add(request.user)
            operation_arrival.save()
            return redirect('/')

        return render(request, 'operation_arrival.html', {'piles': piles, 'brigades': brigades})
    else:
        return redirect('/')

def operation_departure_view(request):
    if request.user.userprofile.role.name == "Производство":
        if request.method == 'POST':
            manager = request.POST.get('manager')
            description = request.POST.get('description')
            number_car_id = request.POST.get('number_car')
            brigade_id = request.POST.get('brigade')  # Получение ID выбранной бригады
            details_list = []

            # Переменная для проверки возможности списания
            can_depart = True
            error_message = ''

            # Итерация по всем полученным данным формы
            for key, value in request.POST.items():
                if key.startswith('pile_') and value:  # Проверка, принадлежит ли ключ к Pile и не пустое ли значение
                    quantity = int(value)  # Преобразование введенного значения в число
                    pile_id = key.split('_')[1]
                    pile = Pile.objects.get(id=pile_id)

                    if quantity > pile.count:  # Проверка наличия свай на складе
                        can_depart = False
                        error_message += f"На складе недостаточно свай {pile.name.name} размером {pile.size}. "
                        break  # Прерываем цикл, т.к. обнаружена ошибка

                    if can_depart and quantity > 0:  # Проверка, что количество больше нуля и списание возможно
                        pile.count -= quantity
                        pile.save()
                        details_list.append({"pile_id": pile_id, "quantity": quantity})

            if not can_depart:  # Если списание невозможно, возвращаем ошибку
                return HttpResponse(error_message, status=400)

            if not details_list:  # Если список деталей пуст, возвращаем ошибку
                return HttpResponse("Пожалуйста, введите количество для хотя бы одной сваи.", status=400)

            details = json.dumps(details_list)  # Преобразование списка деталей в JSON строку

            operation_departure = OperationDeparture(
                manager=manager,
                description=description,
                details=details,
                number_car=number_car_id,
                brigade=BrigadeWork.objects.get(id=brigade_id) if brigade_id else None,
            )
            operation_departure.save()

            return redirect('jb')  # Перенаправление после сохранения
        else:
            piles = Pile.objects.filter(name__pile_type='жб')  # Получаем все объекты Pile для формы
            cars = Car.objects.all()  # Получаем все объекты Car для выбора
            brigades = BrigadeWork.objects.filter(type='жб')  # Получение всех объектов BrigadeWork для выбора
            return render(request, 'operation_departure.html', {'piles': piles, 'cars': cars, 'brigades': brigades, 'car_type_choices': Car.TYPE_CHOICES,})
    else:
        return redirect('/')



def mbp_statistics(request):
    # Извлекаем все записи из модели MBP
    mbps = MBP.objects.all()

    # Передаем данные в шаблон
    return render(request, 'mbp_statistics.html', {'mbps': mbps})


def warehouse_statistics(request):
    # Фильтрация свай по типу "жб"
    piles = Pile.objects.filter(name__pile_type='жб')
    total_count = sum(pile.count for pile in piles)
    total_defect = sum(pile.defect for pile in piles)

    # Получение уникальных имен свай типа "жб"
    pile_names = piles.values_list('name__name', flat=True).distinct()

    piles_stats = {}
    for name in pile_names:
        # Для каждого имени сваи собираем статистику по размерам, учитывая только сваи типа "жб"
        sizes_stats = piles.filter(name__name=name).values('size').annotate(
            total_count=Sum('count'),
            total_defect=Sum('defect')
        ).order_by('size')

        piles_stats[name] = list(sizes_stats)

    context = {
        'total_count': total_count,
        'total_defect': total_defect,
        'piles_stats': piles_stats
    }

    return render(request, 'warehouse_statistics.html', context)

def warehouseV_statistics(request):
    # Фильтрация свай по типу "жб"
    piles = Pile.objects.filter(name__pile_type='винтовые')
    total_count = sum(pile.count for pile in piles)
    total_defect = sum(pile.defect for pile in piles)

    # Получение уникальных имен свай типа "жб"
    pile_names = piles.values_list('name__name', flat=True).distinct()

    piles_stats = {}
    for name in pile_names:
        # Для каждого имени сваи собираем статистику по размерам, учитывая только сваи типа "жб"
        sizes_stats = piles.filter(name__name=name).values('size').annotate(
            total_count=Sum('count'),
            total_defect=Sum('defect')
        ).order_by('size')

        piles_stats[name] = list(sizes_stats)

    context = {
        'total_count': total_count,
        'total_defect': total_defect,
        'piles_stats': piles_stats
    }

    return render(request, 'warehouse_statistics.html', context)
















def search_by_date_range(request):
    if request.user.userprofile.role.name == "Производство" or request.user.userprofile.role.name == "Производство В":
        form = SearchForm(request.GET or None)

        search_results = {}

        if form.is_valid():
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
            manager = form.cleaned_data.get('manager')
            brigade = form.cleaned_data.get('brigade')
            departures_query = None
            arrivals_query = None
            returns_query = None
            beton_query = None
            wire_query = None
            armature_query = None
            order_query = None





            # Фильтрация по датам
            if date_from and date_to:
                filter_date_range = {'date__range': (date_from, date_to)}
            elif date_from:
                filter_date_range = {'date__gte': date_from}
            elif date_to:
                filter_date_range = {'date__lte': date_to}
            else:
                filter_date_range = {}

            if date_from or date_to:
                departures_query = OperationDeparture.objects.all()
                arrivals_query = OperationArrival.objects.all()
                returns_query = ReturnPile.objects.all()
                beton_query = Beton.objects.all()
                wire_query = Wire.objects.all()
                armature_query = Armature.objects.all()
                order_query = Order.objects.all()

                departures_query = departures_query.filter(**filter_date_range)
                arrivals_query = arrivals_query.filter(**filter_date_range)
                returns_query = returns_query.filter(**filter_date_range)
                beton_query = beton_query.filter(**filter_date_range)
                wire_query = wire_query.filter(**filter_date_range)
                armature_query = armature_query.filter(**filter_date_range)
                order_query = order_query.filter(**filter_date_range)

            if manager:
                if departures_query ==None:
                    departures_query = OperationDeparture.objects.all()
                    returns_query = ReturnPile.objects.all()
                departures_query = departures_query.filter(manager=manager)
                returns_query = returns_query.filter(manager=manager)
            if brigade:
                if departures_query !=None:
                    departures_query = OperationDeparture.objects.all()
                departures_query = departures_query.filter(brigade=brigade)

            if arrivals_query !=None:
                for arrival in arrivals_query:
                    if arrival.details:
                        details = json.loads(arrival.details)
                        arrival.piles_info = []
                        for detail in details:
                            try:
                                pile = Pile.objects.get(id=detail["pile_id"])
                                arrival.piles_info.append({
                                    "pile": pile,
                                    "quantity": detail["quantity"]
                                })
                            except Pile.DoesNotExist:
                                continue
            if departures_query !=None:
                for arrival in departures_query:
                    if arrival.details:
                        details = json.loads(arrival.details)
                        arrival.piles_info = []
                        for detail in details:
                            try:
                                pile = Pile.objects.get(id=detail["pile_id"])
                                arrival.piles_info.append({
                                    "pile": pile,
                                    "quantity": detail["quantity"]
                                })
                            except Pile.DoesNotExist:
                                continue

            search_results = {
                'departures': departures_query,
                'arrivals': arrivals_query,
                'returns': returns_query,
                'beton': beton_query,
                'wire': wire_query,
                'armature': armature_query,
                'order': order_query
            }

        return render(request, 'search_by_date.html', {'form': form, 'search_results': search_results})


def materials_overview(request):
    # Получаем 5 последних записей для каждого типа материала
    last_five_beton = Beton.objects.all().order_by('-date')[:5]
    last_five_wire = Wire.objects.all().order_by('-date')[:5]
    last_five_wire4 = Wire4.objects.all().order_by('-date')[:5]
    last_five_wire6 = Wire6.objects.all().order_by('-date')[:5]
    last_five_armature = Armature.objects.all().order_by('-date')[:5]
    last_five_armature10 = Armature10.objects.all().order_by('-date')[:5]

    # Предполагаем, что у нас есть только один экземпляр WirehouseB
    wirehouseb = WirehouseB.objects.first()

    context = {
        'last_five_beton': last_five_beton,
        'last_five_wire': last_five_wire,
        'last_five_wire4': last_five_wire4,
        'last_five_wire6': last_five_wire6,
        'last_five_armature': last_five_armature,
        'last_five_armature10': last_five_armature10,
        'wirehouseb': wirehouseb,
    }
    return render(request, 'materials_overview.html', context)
def classif(request):
    return render(request,'classes.html')

def classifV(request):
    return render(request,'classes_v.html')

def create_order(request):
    if request.user.userprofile.role.name == "Производство":
        if request.method == "POST":
            piles = Pile.objects.filter(name__pile_type='жб')

            total_beton = 0
            total_wire_3 = 0
            total_wire_4 = 0
            total_wire_6 = 0
            total_armature = 0
            total_armature10 = 0
            piles_info = {}  # Обновленный словарь для хранения информации о сваях

            for pile in piles:
                quantity_plates = request.POST.get(f'pile_{pile.id}', 0)
                if quantity_plates=='':
                    continue
                quantity_plates=int(quantity_plates)
                match = re.match(r'(\d+\.?\d*)x(\d+)', pile.size.replace(' ', ''))
                if match:
                    width, length = map(float, match.groups())
                    if width == 1.5 and length == 3:
                        multiplier = 1
                    elif width == 150:
                        multiplier = 9
                    elif width == 200:
                        multiplier = 7
                    else:
                        multiplier = 0
                else:
                        multiplier = 0
                quantity = quantity_plates * multiplier
                if quantity:
                    quantity = int(quantity)
                    if quantity > 0:  # Добавляем в словарь только если количество больше 0
                        pile_key = f"{pile.name.name} {pile.size}"  # Используем имя и размер для ключа
                        piles_info[pile_key] = quantity
                        total_beton += pile.price_beton * quantity
                        total_wire_3 += pile.price_wire_3 * quantity
                        total_wire_4 += pile.price_wire_4 * quantity
                        total_wire_6 += pile.price_wire_6 * quantity
                        if str(pile.name.name)=='ЖБ ф10' or str(pile.name.name)=='Плита':
                            total_armature10 += pile.price_armature * quantity
                        elif str(pile.name.name)=='ЖБ ф8':
                            total_armature += pile.price_armature * quantity


            # Проверяем, были ли добавлены сваи в заказ
            if piles_info:
                order = Order(
                    piles_info=json.dumps(piles_info),  # Сериализация обновленного словаря в JSON строку
                    total_beton=total_beton,
                    total_wire_3=total_wire_3,
                    total_wire_4=total_wire_4,
                    total_wire_6=total_wire_6,
                    total_armature=total_armature,
                    total_armature10=total_armature10,
                )
                order.save()

                return redirect('jb')
            else:
                # Если сваи не были добавлены, возможно, стоит добавить сообщение об ошибке
                return render(request, 'create_order.html', {'piles': piles, 'error': 'No piles were selected.'})

        else:
            piles = Pile.objects.filter(name__pile_type='жб')
            return render(request, 'create_order.html', {'piles': piles})





def materialsV_overview(request):
    # Получаем 5 последних записей для каждого типа материала

    # Предполагаем, что у нас есть только один экземпляр WirehouseB
    wirehousev = WirehouseV.objects.first()

    context = {

        'wirehousev': wirehousev,
    }
    return render(request, 'materials_v.html', context)







def calculate_lopasti(request):
    # Словарь с количеством лопастей, которые можно изготовить из одного листа
    diameters = {89: 66, 108: 56, 133: 39}

    if request.method == 'POST':
        # Обрабатываем данные формы
        thickness = request.POST['thickness']
        inventory = WirehouseV.objects.first()

        # Определяем последнюю цену за лист в зависимости от толщины
        if thickness == '4':
            last_price_entry = Lists.objects.last()
        else:  # Толщина 5 мм
            last_price_entry = Lists5.objects.last()

        # Проверяем, есть ли последняя запись с ценой
        if not last_price_entry:
            return HttpResponse("Ошибка: нет данных о цене.", status=404)

        # Цена за лист
        price_per_sheet = last_price_entry.price

        for diameter, per_sheet in diameters.items():
            input_name = str(diameter)
            count_to_use = int(request.POST.get(input_name, 0))
            blades_to_add = count_to_use * per_sheet

            if thickness == '4':
                attr_name = f'lopasti{diameter}'
                inventory.lists4 -= count_to_use
            else:  # Толщина 5 мм
                attr_name = f'lopasti{diameter}_5'
                inventory.lists5 -= count_to_use

            current_value = getattr(inventory, attr_name, 0)
            setattr(inventory, attr_name, current_value + blades_to_add)

            # Цена за одну лопасть
            price_per_blade = price_per_sheet / per_sheet

            # Создаем запись для каждой изготовленной лопасти
            Lopasti.objects.create(
                size=str(diameter),
                thickness=thickness,
                count=blades_to_add,
                price=price_per_blade
            )

        inventory.save()
        return redirect('/')  # Замените '/' на имя URL вашей начальной страницы

    elif request.method == 'GET':
        # Загружаем страницу с формой
        context = {
            'diameters': diameters.keys()
        }
        return render(request, 'lopasti_calculation.html', context)


def orders_by_day(request):
    # Получаем все заказы, отсортированные по дате
    orders = Order.objects.all().order_by('-date')

    # Для каждого заказа преобразуем информацию о сваях из JSON
    for order in orders:
        # Преобразование строки JSON в объект Python
        order.piles_info = json.loads(order.piles_info)

    return render(request, 'orders_by_day.html', {'orders': orders})


def calculate_tube_cutting(request):
    if request.method == 'POST':
        tubes_info = []
        details_list = []
        formIndex = 0
        while True:
            tube_diameter_key = f'tube_diameter_{formIndex}'
            tube_thickness_key = f'tube_thickness_{formIndex}'
            thickness_key = f'thickness_{formIndex}'
            if tube_diameter_key in request.POST:
                tube_diameter = request.POST[tube_diameter_key]
                tube_thickness = request.POST.get(tube_thickness_key, '0')
                thickness = request.POST.get(thickness_key, '0')
                cuts = {}
                for i in range(1, 8):  # Предполагаем, что у нас есть 4 возможных разреза
                    cut_key = f'cut_{i}_{formIndex}'
                    cuts[f'cut_{i}'] = request.POST.get(cut_key, '0')
                tube_info = {
                    'tube_diameter': tube_diameter,
                    'tube_thickness': tube_thickness,
                    'thickness': thickness,
                    'cuts': cuts
                }
                tubes_info.append(tube_info)
                formIndex += 1
            else:
                break

        for tube in tubes_info:
            tube_diameter = tube['tube_diameter']
            tube_thickness = tube['tube_thickness']
            thickness = tube['thickness']
            cuts = tube['cuts']

            # Определяем соответствующий NamePile
            name_pile, _ = NamePile.objects.get_or_create(name=f'Св {tube_thickness}')

            cut_mappings = {
                'cut_1': (2500, 5),
                'cut_2': [(2500, 2), (3500, 2)],
                'cut_3': [(4500, 2), (3000, 1)],
                'cut_4': (3000, 4),
                'cut_5': (4000, 3),
                'cut_6': [(5000, 2), (2000, 1)],
                'cut_7': [(5000, 1), (3500, 2)],
            }

            for cut_key, value in cuts.items():
                lengths_and_counts = cut_mappings[cut_key]
                if not isinstance(lengths_and_counts[0], tuple):
                    lengths_and_counts = [lengths_and_counts]

                for length, multiplier in lengths_and_counts:
                    count = int(value) * multiplier
                    size = f'{tube_diameter}x{length}'

                    # Находим или создаем сваю с указанными параметрами
                    pile, created = Pile.objects.get_or_create(
                        name=name_pile,
                        size=size,
                        defaults={'count': count}
                    )

                    if not created:
                        pile.count += count
                        pile.save()
                    pile_price = calculate_pile_price(tube_diameter,length, thickness,count)
                    wirehouse, _ = WirehouseV.objects.get_or_create(id=1)  # Предполагаем, что есть один склад
                    tube_amount_to_subtract = (int(length) // 1000) * count
                    lopasti_amount_to_subtract = int(count)

                    # Вычитаем количество труб
                    if str(tube_diameter) == str(89):
                        wirehouse.tube89 -= tube_amount_to_subtract
                        if str(thickness)==str(4):
                            wirehouse.lopasti89 -= count
                        else:
                            wirehouse.lopasti89_5 -= count
                    elif str(tube_diameter) == str(108):
                        wirehouse.tube108 -= tube_amount_to_subtract
                        if str(thickness)==str(4):
                            wirehouse.lopasti108 -= count
                        else:
                            wirehouse.lopasti108_5 -= count
                    elif str(tube_diameter) == str(133):
                        wirehouse.tube133 -= tube_amount_to_subtract
                        if str(thickness)==str(4):
                            wirehouse.lopasti133 -= count
                        else:
                            wirehouse.lopasti133_5 -= count

                    wirehouse.save()
                    if count!=0:
                        details_list.append({
                            'pile_id': pile.id,
                            'quantity': count,
                            'pile_price': str(pile_price)
                        })
        d=OperationArrival_V.objects.create(
            details=json.dumps(details_list),
        )
        d.user.add(request.user)
        d.save()
        return redirect('/')
    else:
        return render(request, 'cut_tube.html')

from decimal import Decimal
def calculate_pile_price(diameter, length, thickness, count):
    # Определение модели трубы в зависимости от диаметра
    tube_model = None
    if diameter == 89:
        tube_model = Tube.objects.all()
    elif diameter == 108:
        tube_model = Tube108.objects.all()
    elif diameter == 133:
        tube_model = Tube133.objects.all()

    if not tube_model:
        return Decimal("0.0")

    last_tube = tube_model.order_by('date').last()

    # Расчет стоимости трубы за длину
    if last_tube:
        tube_price_per_meter = last_tube.unit_price if last_tube.unit_price else Decimal("0.0")
        total_tube_price = tube_price_per_meter * (int(length) // 1000) * count
    else:
        total_tube_price = Decimal("0.0")

    # Получение последней лопасти соответствующего размера и толщины
    last_lopasti = Lopasti.objects.filter(size=str(diameter), thickness=thickness).order_by('date').last()
    lopasti_price = last_lopasti.price * count if last_lopasti else Decimal("0.0")

    # Обновление количества материалов на складе


    # Возвращаем общую стоимость
    total_price = total_tube_price + lopasti_price
    return total_price.quantize(Decimal(".01"))




def operation_arrival_list(request):
    operations = OperationArrival.objects.filter(confirm_b=False)
    for operation in operations:
        details = json.loads(operation.details) if operation.details else []
        operation.piles_info = []  # Добавляем атрибут для хранения информации о сваях

        for detail in details:
            try:
                pile = Pile.objects.get(id=detail["pile_id"])
                operation.piles_info.append({
                    "name": pile.name.name,
                    "size": pile.size,
                    "quantity": detail["quantity"]
                })
            except Pile.DoesNotExist:
                continue  # Пропускаем, если свая не найдена
    if request.method == 'POST':
        operation_id = request.POST.get('operation_id')
        operation = OperationArrival.objects.get(id=operation_id)
        operation.confirm_b = True
        operation.save()
        return redirect('operation_arrival_list')
    return render(request, 'operation_arrival_list.html', {'operations': operations})




#Подтверждения
def confirm_operations(request):
    if request.method == 'POST':
        operation_id = request.POST.get('operation_id')
        action = request.POST.get('action')  # Убедитесь, что action передается правильно

        operation = OperationDeparture.objects.get(id=operation_id)

        if action == 'confirm':
            operation.confirm_b = True
        elif action == 'cancel':
            operation.confirm_b = False
        elif action == 'vis':
            operation.visible = False
            details = json.loads(operation.details)
            for detail in details:
                pile = Pile.objects.get(id=detail["pile_id"])
                pile.count+=int(detail["quantity"])
                pile.save()
        operation.save()
        return redirect('confirm_operations')

    operations_to_confirm = OperationDeparture.objects.filter(confirm_b=False,visible=True)
    confirmed_operations = OperationDeparture.objects.filter(confirm_b=True,visible=True)

    for departure in operations_to_confirm:
                if departure.details:
                    details = json.loads(departure.details)
                    departure.piles_info = []
                    for detail in details:
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            departure.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                            })
                        except Pile.DoesNotExist:
                            continue
    for departure in confirmed_operations:
                if departure.details:
                    details = json.loads(departure.details)
                    departure.piles_info = []
                    for detail in details:
                        try:
                            pile = Pile.objects.get(id=detail["pile_id"])
                            departure.piles_info.append({
                                "pile": pile,
                                "quantity": detail["quantity"]
                            })
                        except Pile.DoesNotExist:
                            continue
    context = {
        'operations_to_confirm': operations_to_confirm,
        'confirmed_operations': confirmed_operations
    }
    return render(request, 'confirm_operations.html', context)






#Добавление
def add_beton(request):
    if request.user.userprofile.role.name == "Производство":
        if request.method == "POST":
            form = BetonForm(request.POST)
            if form.is_valid():
                new_beton = form.save(commit=False)
                wirehouse, _ = WirehouseB.objects.get_or_create(id=1)  # Пример с одним складом
                wirehouse.beton += new_beton.count  # Предполагаем, что арматура используется для бетона
                wirehouse.save()
                new_beton.save()
                return redirect('classes')
        else:
            form = BetonForm()
        return render(request, 'add_beton.html', {'form': form})
    else:
        return redirect('/')
def add_wire(request):
    if request.user.userprofile.role.name == "Производство":
        if request.method == "POST":
            form = WireForm(request.POST)
            if form.is_valid():
                new_wire = form.save(commit=False)
                wirehouse, _ = WirehouseB.objects.get_or_create(id=1)
                wirehouse.wire3 += new_wire.count
                wirehouse.save()
                new_wire.save()
                return redirect('classes')
        else:
            form = WireForm()
        return render(request, 'add_wire.html', {'form': form,'size':'3мм'})
    else:
        return redirect('/')
def add_wire4(request):
    if request.user.userprofile.role.name == "Производство":
        if request.method == "POST":
            form = WireForm4(request.POST)
            if form.is_valid():
                new_wire = form.save(commit=False)
                wirehouse, _ = WirehouseB.objects.get_or_create(id=1)
                wirehouse.wire4 += new_wire.count
                wirehouse.save()
                new_wire.save()
                return redirect('classes')
        else:
            form = WireForm()
        return render(request, 'add_wire.html', {'form': form,'size':'4мм'})
    else:
        return redirect('/')
def add_wire6(request):
    if request.user.userprofile.role.name == "Производство":
        if request.method == "POST":
            form = WireForm6(request.POST)
            if form.is_valid():
                new_wire = form.save(commit=False)
                wirehouse, _ = WirehouseB.objects.get_or_create(id=1)
                wirehouse.wire6 += new_wire.count
                wirehouse.save()
                new_wire.save()
                return redirect('classes')
        else:
            form = WireForm()
        return render(request, 'add_wire.html', {'form': form,'size':'3мм'})
    else:
        return redirect('/')
def add_armature(request):
    if request.user.userprofile.role.name == "Производство":
        if request.method == "POST":
            form = ArmatureForm(request.POST)
            if form.is_valid():
                new_arm = form.save(commit=False)
                wirehouse, _ = WirehouseB.objects.get_or_create(id=1)  # Пример с одним складом
                wirehouse.armature += new_arm.count  # Предполагаем, что арматура используется для бетона
                wirehouse.save()
                new_arm.save()
                return redirect('classes')
        else:
            form = ArmatureForm()
        return render(request, 'add_armature.html', {'form': form,'size':'8мм'})
    else:
        return redirect('/')

def add_armature10(request):
    if request.user.userprofile.role.name == "Производство":
        if request.method == "POST":
            form = Armature10Form(request.POST)
            if form.is_valid():
                new_arm = form.save(commit=False)
                wirehouse, _ = WirehouseB.objects.get_or_create(id=1)
                wirehouse.armature10 += new_arm.count
                wirehouse.save()
                new_arm.save()
                return redirect('classes')
        else:
            form = Armature10Form()
        return render(request, 'add_armature.html', {'form': form,'size':'10мм'})
    else:
        return redirect('/')
def add_tube(request):
    form = TubeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        tube_data = form.save(commit=False)

        # Переводим кг в метры, предполагая, что масса метра трубы d=108, t=4 составляет 10.259 кг
        mass_per_meter = 8.384  # масса одного метра трубы в кг
        meters = round(tube_data.count / mass_per_meter)

        # Обновляем количество метров на складе
        warehouse, _ = WirehouseV.objects.get_or_create(id=1)  # предполагаем, что склад один и его id=1
        warehouse.tube89 = F('tube89') + meters
        warehouse.save()

        # Вычисляем цену за метр
        tube_data.unit_price = tube_data.price / meters if meters else 0
        tube_data.save()

        return redirect('/')  # URL, куда перенаправить пользователя после успешного добавления

    return render(request, 'add_tube.html', {'form': form})

def add_tube108(request):
    form = TubeForm108(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        tube_data = form.save(commit=False)

        # Переводим кг в метры, предполагая, что масса метра трубы d=108, t=4 составляет 10.259 кг
        mass_per_meter = 10.259  # масса одного метра трубы в кг
        meters = round(tube_data.count / mass_per_meter)

        # Обновляем количество метров на складе
        warehouse, _ = WirehouseV.objects.get_or_create(id=1)  # предполагаем, что склад один и его id=1
        warehouse.tube108 = F('tube108') + meters
        warehouse.save()

        # Вычисляем цену за метр
        tube_data.unit_price = tube_data.price / meters if meters else 0
        tube_data.save()

        return redirect('/')  # URL, куда перенаправить пользователя после успешного добавления

    return render(request, 'add_tube.html', {'form': form})

def add_tube108_2(request):
    form = TubeForm108_2(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        tube_data = form.save(commit=False)

        # Переводим кг в метры, предполагая, что масса метра трубы d=108, t=4 составляет 10.259 кг
        mass_per_meter = 9.019  # масса одного метра трубы в кг
        meters = round(tube_data.count / mass_per_meter)

        # Обновляем количество метров на складе
        warehouse, _ = WirehouseV.objects.get_or_create(id=1)  # предполагаем, что склад один и его id=1
        warehouse.tube108_2 = F('tube108_2') + meters
        warehouse.save()

        # Вычисляем цену за метр
        tube_data.unit_price = tube_data.price / meters if meters else 0
        tube_data.save()

        return redirect('/')  # URL, куда перенаправить пользователя после успешного добавления

    return render(request, 'add_tube.html', {'form': form})

def add_tube133(request):
    form = TubeForm133(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        tube_data = form.save(commit=False)

        # Переводим кг в метры, предполагая, что масса метра трубы d=108, t=4 составляет 10.259 кг
        mass_per_meter = 12.725   # масса одного метра трубы в кг
        meters = round(tube_data.count / mass_per_meter)

        # Обновляем количество метров на складе
        warehouse, _ = WirehouseV.objects.get_or_create(id=1)  # предполагаем, что склад один и его id=1
        warehouse.tube133 = F('tube133') + meters
        warehouse.save()

        # Вычисляем цену за метр
        tube_data.unit_price = tube_data.price / meters if meters else 0
        tube_data.save()

        return redirect('/')  # URL, куда перенаправить пользователя после успешного добавления

    return render(request, 'add_tube.html', {'form': form})
def add_lists(request):
    if request.method == "POST":
        thickness = request.POST.get('thickness')
        count = request.POST.get('count')
        company_sell = request.POST.get('company_sell')
        company_buy = request.POST.get('company_buy')
        price = request.POST.get('price')

        if thickness == '4':
            new_list = Lists(count=count, company_sell=company_sell, company_buy=company_buy, price=price)
            wirehouse = WirehouseV.objects.first()  # Предположим, что у вас есть только одна запись
            wirehouse.lists4 += int(count)
        else:
            new_list = Lists5(count=count, company_sell=company_sell, company_buy=company_buy, price=price)
            wirehouse = WirehouseV.objects.first()
            wirehouse.lists5 += int(count)

        new_list.save()
        wirehouse.save()
        return redirect('/vint')  # Перенаправление после сохранения

    return render(request, 'add_list.html')

def add_pile(request):
    if request.method == 'POST':
        form = NamePileAndPileForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('operation_arrival')
    else:
        form = NamePileAndPileForm()

    return render(request, 'add_pile.html', {'form': form})

def update_mbp_view(request):
    if request.method == 'POST':
        brigade_name = request.POST.get('brigade_name', '')  # Получение имени бригады из формы
        details_list = []  # Список для хранения деталей об обновлении количества

        # Обработка количества материалов
        for key, value in request.POST.items():
            if key.startswith('quantity_'):
                mbp_id = key.split('_')[1]
                try:
                    quantity = int(value)
                    if quantity > 0:  # Проверяем, чтобы количество было больше нуля
                        mbp = MBP.objects.get(id=mbp_id)
                        mbp.quantity -= quantity
                        mbp.save()
                        # Добавляем информацию в список деталей
                        details_list.append(f"{mbp.name}: {quantity}")
                except (ValueError, MBP.DoesNotExist):
                    continue  # Пропускаем итерацию при ошибках

        if details_list:  # Если есть детали для сохранения
            details_text = ", ".join(details_list)  # Преобразование списка в строку
            MBP_stat.objects.create(
                admission=True,
                date=timezone.now(),
                details=details_text,
                brigade=brigade_name
            )

        return redirect('/')  # Перенаправление после обработки

    mbps = MBP.objects.all()  # Получение всех объектов MBP для отображения в форме
    return render(request, 'add_mbp.html', {'mbps': mbps})



def add_mbp_view(request):
    if request.method == 'POST':
        details_list = []  # Список для хранения деталей об обновлении количества
        # Обработка количества материалов
        for key, value in request.POST.items():
            if key.startswith('quantity_'):
                mbp_id = key.split('_')[1]
                try:
                    quantity = int(value)
                    if quantity > 0:  # Проверяем, чтобы количество было больше нуля
                        mbp = MBP.objects.get(id=mbp_id)
                        mbp.quantity += quantity
                        mbp.save()
                        # Добавляем информацию в список деталей
                        details_list.append(f"{mbp.name}: {quantity}")
                except (ValueError, MBP.DoesNotExist):
                    continue  # Пропускаем итерацию при ошибках

        if details_list:  # Если есть детали для сохранения
            details_text = ", ".join(details_list)  # Преобразование списка в строку
            MBP_stat.objects.create(
                admission=False,
                date=timezone.now(),
                details=details_text,
            )

        return redirect('/')  # Перенаправление после обработки

    mbps = MBP.objects.all()  # Получение всех объектов MBP для отображения в форме
    return render(request, 'up_mbp.html', {'mbps': mbps})


#Рандомная фигня
def ware(request):
    return render(request, 'ware.html')

def mat(request):
    return render(request, 'mat.html')


def load_car_weight(request):
    car_id = request.GET.get('car_id')
    try:
        if car_id:
            car = Car.objects.get(id=car_id)
            weight = car.weight  # Предполагаем, что weight хранится в тоннах
            return JsonResponse({'weight': weight})
        else:
            return JsonResponse({'error': 'Car ID not provided'}, status=400)
    except Car.DoesNotExist:
        return JsonResponse({'error': 'Car not found'}, status=404)
def load_pile_weight(request):
    pile_id = request.GET.get('pile_id')
    try:
        if pile_id:
            pile = Pile.objects.get(id=pile_id)
            weight = pile.weight
            return JsonResponse({'weight': weight})
        else:
            # Если pile_id не предоставлен, возвращаем вес равный нулю.
            return JsonResponse({'weight': 0})
    except Pile.DoesNotExist:
        # Если свая не найдена, возвращаем ошибку или вес равный нулю.
        return JsonResponse({'error': 'Pile not found'}, status=404)
def load_brigades(request):
    type = request.GET.get('type')
    brigades = BrigadeWork.objects.filter(type=type).order_by('name')
    return JsonResponse(list(brigades.values('id', 'name')), safe=False)
def get_car_numbers(request):
    model = request.GET.get('model')
    numbers = list(Car.objects.filter(model=model).values_list('number', flat=True))
    return JsonResponse({'numbers': numbers})


from collections import defaultdict
from django.http import HttpResponse
import json
import pandas as pd

def export_operation_departures_to_excel(request):
    form = ExportForm(request.POST or None)
    if form.is_valid():
        date_from = form.cleaned_data['date_from']
        date_to = form.cleaned_data['date_to']
        date_to += timedelta(days=1)  # Добавляем один день для включения операций в конечную дату
        brigade = form.cleaned_data['brigade']

        query = OperationDeparture.objects.filter(date__range=[date_from, date_to - timedelta(seconds=1)],confirm_b=True,visible=True)
        if brigade:
            query = query.filter(brigade=brigade)

        data = []
        total_count = 0
        count_by_type = defaultdict(int)

        for operation in query:
            details = json.loads(operation.details)
            for detail in details:
                pile = Pile.objects.get(pk=detail["pile_id"])
                quantity = detail["quantity"]
                total_count += quantity
                count_by_type[pile.name.name] += quantity
                data.append({
                    "Дата": operation.date.strftime("%Y-%m-%d"),
                    "Бригада": operation.brigade.name if operation.brigade else "Не указана",
                    "Менеджер": operation.manager,
                    "Описание": operation.description,
                    "Название сваи": pile.name.name,
                    "Размер сваи": pile.size,
                    "Количество": quantity,
                })

        df = pd.DataFrame(data)

        # Подготовка итоговых данных
        summary_data = [{'Дата': 'Итого:', 'Бригада': '', 'Лид': '', 'Описание': '', 'Название сваи': 'Всего свай', 'Размер сваи': '', 'Количество': total_count}]
        for pile_type, count in count_by_type.items():
            summary_data.append({'Дата': '', 'Бригада': '', 'Лид': '', 'Описание': '', 'Название сваи': f'Итого {pile_type}', 'Размер сваи': '', 'Количество': count})

        summary_df = pd.DataFrame(summary_data)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Отгрузки_{date_from.strftime('%Y-%m-%d')}_{(date_to - timedelta(days=1)).strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Запись данных в Excel
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Отгрузки', index=False)
            summary_df.to_excel(writer, sheet_name='Итоги', index=False, startrow=len(df) + 2)

        return response

    return render(request, 'export_form.html', {'form': form})


def export_operation_arrivals_to_excel(request):
    form = ExportArrForm(request.GET or None)
    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        date_to = date_to + timedelta(days=1)
        brigade = form.cleaned_data.get('brigade')
        operations = OperationArrival.objects.filter(date__range=[date_from, date_to - timedelta(seconds=1)])
        if brigade:
            operations = operations.filter(brigade=brigade)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Operations Data'
        headers = ['Дата', 'Бригада', 'Наименование', 'Размер', 'Количество','МБП','Работа', 'Цена', 'Себестоимость']
        sheet.append(headers)
        total_price = 0
        piles_count = defaultdict(int)
        for operation in operations:
            details = json.loads(operation.details)
            price_parts = json.loads(operation.price_part)

            for detail, price_part in zip(details, price_parts):
                # Предполагаем, что pile_id совпадают и упорядочены одинаково в details и price_part
                if detail['pile_id'] == price_part['pile_id']:
                    # Здесь должна быть логика для получения имени и размера кучи по pile_id
                    # Например, использование словаря или запрос к базе данных
                    pile=Pile.objects.get(id=int(detail['pile_id']))
                    pile_name = pile.name.name
                    pile_size = pile.size

                    quantity = detail['quantity']
                    mbps=11.28 * quantity

                    if pile.name=='Плита':
                        price = float(price_part['part_price'])+500*int(quantity)+11.28*int(quantity)
                        works=500 * quantity
                    else:
                        match = re.match(r'(\d+\.?\d*)x(\d+)', pile.size.replace(' ', ''))
                        width, length = map(float, match.groups())
                        if int(length)<4999:
                            price = round(float(price_part['part_price']) + 115 * quantity+11.28 * quantity,3)
                            works=115 * quantity
                        else:
                            price = round(float(price_part['part_price']) + 230 * quantity+11.28 * quantity,3)
                            works=230 * quantity
                    price1 = float(price_part['part_price'])
                    cost_price = float(price1) / quantity if quantity != 0 else 0
                    total_price += price
                    pile_key = f"{pile_name} {pile_size}"
                    piles_count[pile_key] += quantity
                    row = [
                        operation.date.strftime("%Y-%m-%d"),
                        operation.brigade.name if operation.brigade else '',
                        pile_name,
                        pile_size,
                        quantity,
                        mbps,
                        works,
                        price,
                        cost_price
                    ]

                    sheet.append(row)
        sheet.append(["", "", "", "Итого:", "", total_price, ""])
        for pile_info, qty in piles_count.items():
            sheet.append(["", "", f"{pile_info}:", f"{qty} шт", "", "", ""])
        filename = 'Operations_Export.xlsx'
        response = HttpResponse(content_type='application/ms-excel')
        content = f'attachment; filename="{filename}"'
        response['Content-Disposition'] = content
        workbook.save(response)

        return response
    else:
        return render(request, 'export_arr_form.html', {'form': form})



def order_tube_cutting(request):
    if request.method == 'POST':
        tubes_info = []
        piles_list = []
        t_list = {'Труба d89':0,'Труба d108':0,'Труба d133':0}
        lopasti_list={'Лопасть d89 4мм':0,'Лопасть d108 4мм':0,'Лопасть d133 4мм':0,'Лопасть d89 5мм':0,'Лопасть d108 5мм':0,'Лопасть d133 5мм':0}
        formIndex = 0
        while True:
            tube_diameter_key = f'tube_diameter_{formIndex}'
            tube_thickness_key = f'tube_thickness_{formIndex}'
            thickness_key = f'thickness_{formIndex}'
            if tube_diameter_key in request.POST:
                tube_diameter = request.POST[tube_diameter_key]
                tube_thickness = request.POST.get(tube_thickness_key, '0')
                thickness = request.POST.get(thickness_key, '0')
                cuts = {}
                for i in range(1, 8):  # Предполагаем, что у нас есть 4 возможных разреза
                    cut_key = f'cut_{i}_{formIndex}'
                    cuts[f'cut_{i}'] = request.POST.get(cut_key, '0')
                tube_info = {
                    'tube_diameter': tube_diameter,
                    'tube_thickness': tube_thickness,
                    'thickness': thickness,
                    'cuts': cuts
                }
                tubes_info.append(tube_info)
                formIndex += 1
            else:
                break

        for tube in tubes_info:
            tube_diameter = tube['tube_diameter']
            tube_thickness = tube['tube_thickness']
            thickness = tube['thickness']
            cuts = tube['cuts']

            # Определяем соответствующий NamePile
            name_pile, _ = NamePile.objects.get_or_create(name=f'Св {tube_thickness}')

            cut_mappings = {
                'cut_1': (2500, 5),
                'cut_2': [(2500, 2), (3500, 2)],
                'cut_3': [(4500, 2), (3000, 1)],
                'cut_4': (3000, 4),
                'cut_5': (4000, 3),
                'cut_6': [(5000, 2), (2000, 1)],
                'cut_7': [(5000, 1), (3500, 2)],
            }

            for cut_key, value in cuts.items():
                lengths_and_counts = cut_mappings[cut_key]
                if not isinstance(lengths_and_counts[0], tuple):
                    lengths_and_counts = [lengths_and_counts]

                for length, multiplier in lengths_and_counts:
                    count = int(value) * multiplier
                    size = f'{tube_diameter}x{length}'

                    # Находим или создаем сваю с указанными параметрами
                    pile, created = Pile.objects.get_or_create(
                        name=name_pile,
                        size=size,
                        defaults={'count': count}
                    )

                    if count!=0:
                        piles_list.append({
                            'pile_id': pile.id,
                            'quantity': count,
                        })
                        t_list[f'Труба d{tube_diameter}']+=(int(length)//1000)*int(count)
                        lopasti_list[f'Лопасть d{tube_diameter} {thickness}мм']+=int(count)
        d=Order_v.objects.create(
            piles=json.dumps(piles_list),
            tubes=json.dumps(t_list),
            lopastis=json.dumps(lopasti_list),
        )
        d.save()
        return redirect('/')
    else:
        return render(request, 'order_tube.html')



def download_orders(request):
    if request.method == 'POST':
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')

        # Преобразование строк в объекты datetime
        date_from = datetime.strptime(date_from, '%Y-%m-%d')
        date_to = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="orders_{date_from}_{date_to}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

        # Заголовки столбцов
        headers = ["Дата", "бетон", "проволка 3мм", "проволка 4мм", "проволка 6мм", "арматура 8мм","арматура 10мм"]
        for col_num, header in enumerate(headers, 1):
            ws[get_column_letter(col_num) + '1'] = header

        # Получение данных заказов
        orders = Order.objects.filter(date__range=[date_from, date_to - timedelta(seconds=1)])
        row_num = 2
        total_beton = total_wire_3 = total_wire_4 = total_wire_6 = total_armature= total_armature10 = Decimal('0.0')

        for order in orders:
            ws[f'A{row_num}'] = order.date.strftime('%Y-%m-%d')
            ws[f'B{row_num}'] = str(order.total_beton)
            ws[f'C{row_num}'] = str(order.total_wire_3)
            ws[f'D{row_num}'] = str(order.total_wire_4)
            ws[f'E{row_num}'] = str(order.total_wire_6)
            ws[f'F{row_num}'] = str(order.total_armature)
            ws[f'G{row_num}'] = str(order.total_armature10)
            total_beton += order.total_beton
            total_wire_3 += order.total_wire_3
            total_wire_4 += order.total_wire_4
            total_wire_6 += order.total_wire_6
            total_armature += order.total_armature
            total_armature10 += order.total_armature10

            row_num += 1

        # Добавление строки с общими суммами
        ws[f'A{row_num}'] = 'Всего'
        ws[f'B{row_num}'] = str(total_beton)
        ws[f'C{row_num}'] = str(total_wire_3)
        ws[f'D{row_num}'] = str(total_wire_4)
        ws[f'E{row_num}'] = str(total_wire_6)
        ws[f'F{row_num}'] = str(total_armature)
        ws[f'G{row_num}'] = str(total_armature10)
        row_num += 2  # Добавляем пробел перед деталями заказов

        # Добавление заголовков для деталей заказов
        ws[f'A{row_num}'] = 'Детали заказов'
        row_num += 1
        ws[f'A{row_num}'] = 'Дата'
        ws[f'B{row_num}'] = 'Тип сваи'
        ws[f'C{row_num}'] = 'Количество'
        row_num += 1

        for order in orders:
            piles_info = json.loads(order.piles_info)
            for pile_info in piles_info.items():
                ws[f'A{row_num}'] = order.date.strftime('%Y-%m-%d')
                ws[f'B{row_num}'] = pile_info[0]  # Имя сваи
                ws[f'C{row_num}'] = pile_info[1]  # Количество
                row_num += 1

        wb.save(response)
        return response
    else:
        return render(request, 'download_orders.html')

def export_material_data(request):
    form = MaterialExportForm(request.GET or None)
    if request.method == 'GET' and form.is_valid():
        material_type = form.cleaned_data['material']
        date_from = form.cleaned_data['date_from']
        date_to = form.cleaned_data['date_to']

        # Устанавливаем date_to в конец дня
        date_to = date_to + timedelta(days=1)

        model_map = {
            'beton': Beton,
            'wire': Wire,
            'wire4': Wire4,
            'wire6': Wire6,
            'armature': Armature,
            'armature10': Armature10
        }

        model = model_map[material_type]
        records = model.objects.filter(date__range=[date_from, date_to - timedelta(seconds=1)])  # убавляем секунду, чтобы остаться в пределах суток

        # Создание Excel файла
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={material_type}_data_{date_from}_{date_to}.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = 'Material Data'

        # Заголовки для Excel
        columns = ['Date', 'Company Sell', 'Company Buy', 'Count', 'Price']
        for i, column in enumerate(columns, 1):
            ws[get_column_letter(i) + '1'] = column

        # Добавление данных в Excel
        for row, record in enumerate(records, start=2):
            ws[f'A{row}'] = record.date.strftime("%Y-%m-%d")
            ws[f'B{row}'] = record.company_sell
            ws[f'C{row}'] = record.company_buy
            ws[f'D{row}'] = record.count
            ws[f'E{row}'] = float(record.price)

        wb.save(response)
        return response

    # Показываем форму, если GET запрос без параметров или форма не валидна
    return render(request, 'export_material_form.html', {'form': form})


def download_excel(request):
    return render(request, 'download_excel.html')



def display_mbp_stats(request):
    # Получение 10 последних записей где admission = True
    mbp_stats_true = MBP_stat.objects.filter(admission=True).order_by('-date')[:10]

    # Получение 10 последних записей где admission = False
    mbp_stats_false = MBP_stat.objects.filter(admission=False).order_by('-date')[:10]

    # Функция для преобразования строки деталей в словарь
    def parse_details(details):
        details_dict = {}
        items = details.split(", ")
        for item in items:
            if ': ' in item:
                key, value = item.split(': ')
                details_dict[key.strip()] = int(value.strip())
        return details_dict

    # Расширяем объекты stat, добавляя преобразованные детали
    for stat in mbp_stats_true:
        stat.details_dict = parse_details(stat.details)

    for stat in mbp_stats_false:
        stat.details_dict = parse_details(stat.details)

    context = {
        'mbp_stats_true': mbp_stats_true,
        'mbp_stats_false': mbp_stats_false
    }

    return render(request, 'mbp_stats.html', context)

def latest_deliveries(request):
    # Получаем последние пять операций списания материалов
    last_five_dels = Del_mat.objects.order_by('-date')

    # Передаем их в контекст для использования в шаблоне
    context = {
        'deliveries': last_five_dels
    }
    return render(request, 'latest_deliveries.html', context)

from django.http import JsonResponse

def get_brigades(request):
    query = request.GET.get('query', '')
    print('----------------------------')
    print("Query received:", query)
    brigades = BrigadeWork.objects.filter(name__icontains=query)
    results = [{'id': brigade.id, 'name': brigade.name} for brigade in brigades]
    print("Results:", results)
    return JsonResponse(results, safe=False)


from django.shortcuts import render, redirect
from .models import Debt, Pile, BrigadeWork
from django.contrib import messages
from django.views import View

class AddDebtView(View):
    def get(self, request):
        brigades = BrigadeWork.objects.all()
        piles = Pile.objects.filter(name__pile_type='жб')
        return render(request, 'dept_form.html', {'brigades': brigades, 'piles': piles})

    def post(self, request):
        brigade_id = request.POST.get('brigade')
        brigade = get_object_or_404(BrigadeWork, id=brigade_id)
        debt = get_object_or_404(Debt, brigade=brigade)

        details = json.loads(debt.details)
        added_piles = {}

        for key, value in request.POST.items():
            if key.startswith('pile_'):
                pile_id = key.split('_')[1]
                quantity = int(value)
                if quantity > 0:
                    pile = get_object_or_404(Pile, id=int(pile_id))
                    pile.count -= quantity
                    pile.save()
                    if pile_id in details:
                        details[pile_id] += quantity
                    else:
                        details[pile_id] = quantity
                    added_piles[pile_id] = quantity

        debt.details = json.dumps(details)
        debt.save()

        # Сохранение в Detail_Debt только добавленных сваев
        if added_piles:
            detail_debt = Detail_Debt(
                details=json.dumps(added_piles),
                brigade=brigade,
                confirm=False,
                date=timezone.now()
            )
            detail_debt.save()

        return redirect('/')


def display_debts(request):
    debts = Debt.objects.all()# Получаем все долги, отсортированные по дате

    # Дополнение каждого долга информацией о сваях
    for debt in debts:
        details = json.loads(debt.details)
        piles_info = []
        for item in details:
            pile = Pile.objects.get(id=item["pile_id"])  # Получаем сваю по ID
            piles_info.append({
                "name": pile.name.name,
                "size": pile.size,
                "quantity": item["quantity"]
            })
        debt.piles_info = piles_info  # Добавляем информацию о сваях к объекту долга

    return render(request, 'display_debts.html', {'debts': debts})
#{"pile_id": pile_id, "quantity": quantity}
import re
def process_text_view(request):
    pile_pattern = re.compile(r'(\d{1,3})[хx\-](\d{4})[хx\-](\d{3})(-?\d*мм)?(?=\D|$)') # Для извлечения размеров сваи
    order_pattern = re.compile(r'^([\w\d/\-]+)')
    rebar_pattern = re.compile(r'\b(\d+) ?мм\b')

    def extract_order_number(text):
        order_match = order_pattern.match(text)
        return order_match.group(1) if order_match else "Не найден"

    if request.method == 'POST':
        if 'confirm' in request.POST:
            details = request.session.get('details')
            order_match = request.session.get('order_match')
            brigade_id = request.session.get('brigade_id')
            rebar_size = request.session.get('rebar_size')
            piles = json.loads(details)
            brigade = BrigadeWork.objects.get(id=brigade_id)

            operation_departure = OperationDeparture(
                manager=order_match,
                details=details,
                number_car='12333',
                brigade=brigade,
            )
            operation_departure.save()

            op = OperationDepCount(
                details=details,
                operation=operation_departure
            )
            send = SendOperation(
                details=details,
                operation=operation_departure
            )
            send.save()
            op.save()

            return redirect('/')
        else:
            form = TextProcessingForm(request.POST)
            if form.is_valid():
                text = form.cleaned_data['text_input']
                brigade = form.cleaned_data['brigade']
                pile_match = pile_pattern.findall(text)
                if pile_match:
                    sizes = pile_match  # Получаем размеры сваи
                    modified_text = pile_pattern.sub('', text)
                else:
                    sizes = (None, None, None)
                    modified_text = text

                remaining_text = modified_text
                order_match = extract_order_number(remaining_text)
                rebar_match = rebar_pattern.search(text)
                rebar_size = int(rebar_match.group(1)) if rebar_match else 10
                rebar = str(rebar_size) # Если нет совпадения, по умолчанию ставим 10
                pile_name = f"ЖБ ф{rebar}"
                name_pile = NamePile.objects.filter(name=pile_name).first()
                piles = []
                for i in sizes:
                    count = int(i[0])
                    length = i[2]
                    width = i[1]
                    size_f = f'{length}x{width}'
                    pile = Pile.objects.get(name=name_pile, size=size_f)
                    piles.append({"pile_id": pile.id, "quantity": count, "name": pile.name.name, "size": pile.size})

                details = json.dumps(piles)
                brigade_id = brigade.id if brigade else None

                # Сохраняем данные в сессии для последующего подтверждения
                request.session['details'] = details
                request.session['order_match'] = order_match
                request.session['brigade_id'] = brigade_id
                request.session['rebar_size'] = rebar_size

                return render(request, 'confirm_operation.html', {'piles': piles, 'order_match': order_match, 'brigade': brigade})

    else:
        form = TextProcessingForm()

    return render(request, 'process_text_form.html', {'form': form})

def confirm_operation_view(request):
    if request.method == 'POST':
        if 'confirm' in request.POST:
            details = request.session.get('details')
            order_match = request.session.get('order_match')
            brigade_id = request.session.get('brigade_id')
            rebar_size = request.session.get('rebar_size')
            piles = json.loads(details)
            brigade = BrigadeWork.objects.get(id=brigade_id)

            operation_departure = OperationDeparture(
                manager=order_match,
                details=details,
                number_car='12333',
                brigade=brigade,
            )
            operation_departure.save()

            op = OperationDepCount(
                details=details,
                operation=operation_departure
            )
            send = SendOperation(
                details=details,
                operation=operation_departure
            )
            send.save()
            op.save()

            return redirect('/')

    return redirect('process_text_view')


def update_pile_count(request, count_id):
    count_entry = get_object_or_404(OperationDepCount, id=count_id)
    piles = json.loads(count_entry.details)

    if request.method == 'POST':
        for pile_detail in piles:
            input_name = f"quantity_{pile_detail['pile_id']}"
            if input_name in request.POST:
                new_quantity = int(request.POST[input_name])
                pile_detail['quantity'] = max(0, pile_detail['quantity'] - new_quantity)

        # Обновляем JSON details
        count_entry.details = json.dumps(piles)
        count_entry.visible = any(pile['quantity'] > 0 for pile in piles)
        count_entry.save()

        return redirect('/')  # Перенаправляем на страницу просмотра

    # Преобразуем ID сваи в наименование для отображения
    for pile in piles:
        pile_object = Pile.objects.get(id=pile['pile_id'])
        pile['name'] = pile_object.name.name
        pile['size'] = pile_object.size

    return render(request, 'update_pile_count.html', {'count_entry': count_entry, 'piles': piles})

def add_pile_to_count(request, count_id):
    count_entry = get_object_or_404(OperationDepCount, id=count_id)
    debt = get_object_or_404(Debt, brigade=count_entry.operation.brigade)
    operation_departure = count_entry.operation

    debt_details = json.loads(debt.details)
    choices = [(pile_id, f"{Pile.objects.get(id=pile_id).name.name} {Pile.objects.get(id=pile_id).size} (Доступно: {quantity})")
               for pile_id, quantity in debt_details.items() if quantity > 0]

    if request.method == 'POST':
        form = AddPileForm(request.POST)
        form.fields['pile'].choices = choices
        if form.is_valid():
            pile_id = form.cleaned_data['pile']
            quantity = form.cleaned_data['quantity']
            pile_object = Pile.objects.get(id=pile_id)

            # Обновляем записи в OperationDepCount
            piles = json.loads(count_entry.details)
            for pile in piles:
                if pile['pile_id'] == int(pile_id):
                    pile['quantity'] += quantity
                    break
            else:
                piles.append({
                    'pile_id': int(pile_id),
                    'quantity': quantity,
                    'name': pile_object.name.name,
                    'size': pile_object.size
                })
            count_entry.details = json.dumps(piles)
            count_entry.save()

            # Обновляем записи в Debt
            debt_details[pile_id] -= quantity
            debt.details = json.dumps(debt_details)
            debt.save()

            # Обновляем поле extra_details в OperationDeparture
            try:
                extra_details = json.loads(operation_departure.extra_details)
            except (json.JSONDecodeError, TypeError):
                extra_details = []

            for pile in extra_details:
                if pile['pile_id'] == int(pile_id):
                    pile['quantity'] += quantity
                    break
            else:
                extra_details.append({
                    'pile_id': int(pile_id),
                    'quantity': quantity,
                    'name': pile_object.name.name,
                    'size': pile_object.size
                })

            operation_departure.extra_details = json.dumps(extra_details)
            operation_departure.save()

            return redirect('update_pile_count', count_id=count_id)
    else:
        form = AddPileForm()
        form.fields['pile'].choices = choices

    return render(request, 'add_pile_to_count.html', {'form': form, 'count_entry': count_entry})


from django.shortcuts import render, redirect
from .models import OperationDepCount, Pile, OperationDeparture
import json

def list_operation_dep_count(request):
    operations = OperationDepCount.objects.filter(confirm=False)

    for operation in operations:
        # Получаем связанную операцию
        linked_operation = OperationDeparture.objects.get(id=operation.operation.id)
        linked_details = json.loads(linked_operation.details)
        extra_details = json.loads(linked_operation.extra_details) if linked_operation.extra_details else []
        all_details = linked_details + extra_details

        # Обогащенные детали из OperationDeparture
        enriched_details = []
        for detail in linked_details:
            pile = Pile.objects.get(id=detail["pile_id"])
            enriched_details.append({
                "name": pile.name.name,
                "size": pile.size,
                "quantity": detail["quantity"]
            })
        operation.enriched_details = enriched_details

        # Детали из OperationDepCount
        dep_count_details = json.loads(operation.details)
        done_details = []
        for detail in dep_count_details:
            pile = Pile.objects.get(id=detail["pile_id"])
            done_details.append({
                "name": pile.name.name,
                "size": pile.size,
                "quantity": detail["quantity"]
            })

        # Вычисляем оставшиеся сваи
        remaining_details = []
        for linked_detail in all_details:
            remaining_quantity = linked_detail["quantity"]
            for dep_count_detail in dep_count_details:
                if linked_detail["pile_id"] == dep_count_detail["pile_id"]:
                    remaining_quantity -= dep_count_detail["quantity"]
            remaining_details.append({
                "name": Pile.objects.get(id=linked_detail["pile_id"]).name.name,
                "size": Pile.objects.get(id=linked_detail["pile_id"]).size,
                "quantity": max(remaining_quantity, 0)
            })
        operation.remaining_details = remaining_details

        operation.done_details = done_details

    if request.method == 'POST':
        count_id = request.POST.get('count_id')
        operation_count = OperationDepCount.objects.get(id=count_id)
        operation_count.confirm = True
        operation_count.save()
        return redirect('list_operation_dep_count')

    context = {
        'operations': operations,
    }
    return render(request, 'list_operation_dep_count.html', context)

class ReturnPilesView(View):
    def get(self, request):
        user_brigade = get_object_or_404(UserBrigade, user=request.user)
        piles = Pile.objects.filter(name__pile_type='жб')
        return render(request, 'return_piles_form.html', {'user_brigade': user_brigade, 'piles': piles})

    def post(self, request):
        user_brigade = get_object_or_404(UserBrigade, user=request.user)
        details = {}
        description = request.POST.get('description', '')

        for key, value in request.POST.items():
            if key.startswith('pile_') and value:
                pile_id = key.split('_')[1]
                quantity = int(value)
                if quantity > 0:
                    details[pile_id] = quantity

        return_piles = ReturnPiles(
            brigade=user_brigade.brigade,
            details=json.dumps(details),
            description=description
        )
        return_piles.save()

        return redirect('/')



class ReturnPilesListView(View):
    def get(self, request):
        return_piles_list = ReturnPiles.objects.filter(confirm_s=False)
        for return_pile in return_piles_list:
            return_pile.details_dict = json.loads(return_pile.details)
            for pile_id, quantity in return_pile.details_dict.items():
                pile = Pile.objects.get(id=pile_id)
                return_pile.details_dict[pile_id] = {
                    'name': pile.name.name,
                    'size': pile.size,
                    'quantity': quantity
                }
        return render(request, 'return_piles_list.html', {'return_piles_list': return_piles_list})

    def post(self, request):
        return_pile_id = request.POST.get('return_pile_id')
        return_pile = get_object_or_404(ReturnPiles, id=return_pile_id)
        return_pile.confirm_s = True
        return_pile.save()
        return redirect('/')  # Перенаправление на список

class ConfirmSkladView(View):
    def get(self, request):
        return_piles_list = ReturnPiles.objects.filter(confirm_s=True, confirm_sklad=False)
        for return_pile in return_piles_list:
            return_pile.details_dict = json.loads(return_pile.details)
            for pile_id, quantity in return_pile.details_dict.items():
                pile = Pile.objects.get(id=int(pile_id))  # Преобразование в целое число
                return_pile.details_dict[pile_id] = {
                    'name': pile.name.name,
                    'size': pile.size,
                    'quantity': quantity
                }
        return render(request, 'return_piles_list.html', {'return_piles_list': return_piles_list})

    def post(self, request):
        return_pile_id = request.POST.get('return_pile_id')
        return_pile = get_object_or_404(ReturnPiles, id=return_pile_id)
        return_pile.confirm_sklad = True
        return_pile.save()
        di = json.loads(return_pile.details)
        for pile_id, quantity in di.items():
            pile = Pile.objects.get(id=int(pile_id))  # Преобразование в целое число
            pile.count += int(quantity)
            pile.save()
        return redirect('/')


class DebtListView(View):
    def get(self, request):
        brigades = BrigadeWork.objects.all()
        return render(request, 'debt_list.html', {'brigades': brigades})

# Представление для редактирования долга конкретной бригады
class DebtDetailView(View):
    def get(self, request, brigade_id):
        debt = get_object_or_404(Debt, brigade_id=brigade_id)
        debt_details = json.loads(debt.details)
        piles = {Pile.objects.get(id=pile_id): quantity for pile_id, quantity in debt_details.items()}
        return render(request, 'debt_detail.html', {'debt': debt, 'piles': piles})

    def post(self, request, brigade_id):
        debt = get_object_or_404(Debt, brigade_id=brigade_id)
        debt_details = json.loads(debt.details)

        for pile_id in debt_details.keys():
            quantity = request.POST.get(f'quantity_{pile_id}', 0)
            debt_details[pile_id] = int(quantity)

        debt.details = json.dumps(debt_details)
        debt.save()
        return redirect('debt_list')

def confirm_departures_view(request):
    send_details = SendDetail.objects.filter(confirm=False)
    detail_debts = Detail_Debt.objects.filter(confirm=False)

    # Расширяем details для отображения в send_details
    for detail in send_details:
        detail.piles_info = []
        try:
            details = json.loads(detail.details)
            for item in details:
                if isinstance(item, dict) and "pile_id" in item and "quantity" in item:
                    pile_id = item["pile_id"]
                    quantity = item["quantity"]
                    if pile_id is not None:
                        pile = Pile.objects.get(id=int(pile_id))
                        detail.piles_info.append({
                            "pile": pile,
                            "quantity": quantity
                        })
                else:
                    print(f"Unexpected item format in send_detail id={detail.id}: {item}")
        except (json.JSONDecodeError, Pile.DoesNotExist) as e:
            detail.piles_info = []
            print(f"Error processing send_detail id={detail.id}: {e}")

    # Расширяем details для отображения в detail_debts
    for debt in detail_debts:
        debt.piles_info = []
        try:
            details = json.loads(debt.details)
            if isinstance(details, dict):
                for pile_id, quantity in details.items():
                    pile = Pile.objects.get(id=int(pile_id))
                    debt.piles_info.append({
                        "pile": pile,
                        "quantity": quantity
                    })
            else:
                print(f"Unexpected details format in detail_debt id={debt.id}: {details}")
        except (json.JSONDecodeError, Pile.DoesNotExist) as e:
            debt.piles_info = []
            print(f"Error processing detail_debt id={debt.id}: {e}")

    if request.method == 'POST':
        detail_id = request.POST.get('detail_id')
        detail_type = request.POST.get('detail_type')

        if detail_type == 'send':
            detail = get_object_or_404(SendDetail, id=detail_id)
        elif detail_type == 'debt':
            detail = get_object_or_404(Detail_Debt, id=detail_id)
        else:
            detail = None

        if detail:
            detail.confirm = True
            detail.save()

        return redirect('confirm_departures')

    return render(request, 'confirm_departures.html', {'send_details': send_details, 'detail_debts': detail_debts})



def update_send_operation(request, operation_id):
    operation = get_object_or_404(OperationDeparture, id=operation_id)
    send_operation = get_object_or_404(SendOperation, operation=operation)
    details = json.loads(send_operation.details)

    if request.method == 'POST':
        for detail in details:
            input_name = f"quantity_{detail['pile_id']}"
            if input_name in request.POST:
                new_quantity = int(request.POST[input_name])
                detail['quantity'] = max(0, detail['quantity'] - new_quantity)

                # Создаем запись в SendDetail
                SendDetail.objects.create(
                    operation=operation,
                    details=json.dumps({"pile_id": detail['pile_id'], "quantity": new_quantity}),
                    confirm=False
                )

        # Обновляем JSON details
        send_operation.details = json.dumps(details)
        send_operation.save()

        # Проверяем, все ли quantity равны 0
        if all(detail['quantity'] == 0 for detail in details):
            operation.confirm_b = True
            operation.save()

        return redirect('/')  # Перенаправляем на нужную страницу

    # Преобразуем ID сваи в наименование для отображения
    for detail in details:
        pile_object = Pile.objects.get(id=detail['pile_id'])
        detail['name'] = pile_object.name.name
        detail['size'] = pile_object.size

    return render(request, 'update_send_operation.html', {'operation': operation, 'details': details})






class ToolOperationView(View):
    def get(self, request):
        form = ToolOperationForm()
        brigades = BrigadeWork.objects.all()
        tools = Tool.objects.all()
        return render(request, 'tool_operation_form.html', {'form': form, 'brigades': brigades, 'tools': tools})

    def post(self, request):
        form = ToolOperationForm(request.POST)
        if form.is_valid():
            brigade = form.cleaned_data['brigade']
            operation = form.cleaned_data['operation']
            tool_details = {}

            for key, value in request.POST.items():
                if key.startswith('tool_') and value:
                    tool_id = int(key.split('_')[1])
                    quantity = int(value)
                    tool = Tool.objects.get(id=tool_id)

                    # Обновляем данные в Tool_details
                    tool_detail_entry = Tool_details(
                        details=json.dumps({tool_id: quantity}),
                        brigade=brigade,
                        operation=operation
                    )
                    tool_detail_entry.save()

                    # Обновляем или создаем данные в BrigadeTool
                    brigade_tool, created = BrigadeTool.objects.get_or_create(brigade=brigade)
                    brigade_tool_details = json.loads(brigade_tool.details) if brigade_tool.details else {}

                    if operation == 'Дать':
                        if str(tool_id) in brigade_tool_details:
                            brigade_tool_details[str(tool_id)] += quantity
                        else:
                            brigade_tool_details[str(tool_id)] = quantity
                    elif operation == 'Списать':
                        if str(tool_id) in brigade_tool_details:
                            brigade_tool_details[str(tool_id)] -= quantity
                            if brigade_tool_details[str(tool_id)] <= 0:
                                del brigade_tool_details[str(tool_id)]
                        else:
                            brigade_tool_details[str(tool_id)] = -quantity

                    brigade_tool.details = json.dumps(brigade_tool_details)
                    brigade_tool.save()

            return redirect('/')

        brigades = BrigadeWork.objects.all()
        tools = Tool.objects.all()
        return render(request, 'tool_operation_form.html', {'form': form, 'brigades': brigades, 'tools': tools})

class ToolOperationListView(View):
    def get(self, request):
        operations = Tool_details.objects.order_by('-date')[:10]

        for operation in operations:
            operation_details = json.loads(operation.details)
            enriched_details = {}
            for tool_id, quantity in operation_details.items():
                tool = get_object_or_404(Tool, id=tool_id)
                enriched_details[tool.name] = quantity
            operation.details = enriched_details

        return render(request, 'tool_operation_list.html', {'operations': operations})

def get_brigade_tools(request, brigade_id):
    brigade_tool = BrigadeTool.objects.filter(brigade_id=brigade_id).first()
    tools = []
    if brigade_tool:
        details = json.loads(brigade_tool.details)
        for tool_id, count in details.items():
            tool = Tool.objects.get(id=tool_id)
            tools.append({'name': tool.name, 'count': count})
    return JsonResponse(tools, safe=False)

class ExportToolDetailsView(View):
    def get(self, request):
        form = DateRangeForm()
        return render(request, 'export_tool_details.html', {'form': form})

    def post(self, request):
        form = DateRangeForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']

            # Фильтруем записи по дате
            if end_date:
                tool_details = Tool_details.objects.filter(date__range=[start_date, end_date])
            else:
                tool_details = Tool_details.objects.filter(date__gte=start_date)

            # Создаем DataFrame для данных
            data = []
            for detail in tool_details:
                details_dict = json.loads(detail.details)
                for tool_id, quantity in details_dict.items():
                    tool = Tool.objects.get(id=tool_id)
                    data.append({
                        'Дата': detail.date.strftime('%Y-%m-%d'),
                        'Бригада': detail.brigade.name,
                        'Операция': detail.operation,
                        'Инструмент': tool.name,
                        'Количество': quantity
                    })

            df = pd.DataFrame(data)

            # Генерируем Excel файл
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=tool_details_{start_date}_to_{end_date if end_date else "present"}.xlsx'

            with pd.ExcelWriter(response, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Tool Details')

            return response

        return render(request, 'export_tool_details.html', {'form': form})
